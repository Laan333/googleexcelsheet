import openpyxl
import datetime
import os
import logging

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s | %(levelname)s | %(message)s',
    datefmt='%d.%m.%Y %H:%M:%S'
)
logger = logging.getLogger("excel_worker")


class ExcelSingleFileWorker:
    def __init__(self, filepath: str, date: datetime.date = None):
        self.filepath = filepath
        self.report_sheet = "Report"
        self.archive_date = date or self._get_previous_month_date()
        self.current_month_sheet = self.archive_date.strftime("%m.%Y")

        if not os.path.exists(self.filepath):
            self._create_initial_workbook()
            logger.info(f"Создан новый Excel-файл и лист '{self.report_sheet}'")

        self.wb = openpyxl.load_workbook(self.filepath)

    def _get_previous_month_date(self) -> datetime.date:
        today = datetime.date.today()
        first_of_this_month = today.replace(day=1)
        last_month = first_of_this_month - datetime.timedelta(days=1)
        return last_month

    def _create_initial_workbook(self):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        wb.create_sheet(self.report_sheet)
        wb.save(self.filepath)

    def _sanitize_cell(self, cell):
        """ Обрабатываем ячейку, преобразуя её в строку или оставляем как есть.
        ПРИМЕЧАНИЕ: Этот метод больше не используется для архивации,
        так как мы копируем данные напрямую с сохранением форматирования. """
        if isinstance(cell, (datetime.datetime, datetime.date)):
            return cell.strftime('%d.%m.%Y')
        elif isinstance(cell, float):  # если число с плавающей точкой
            # Просто возвращаем исходное значение, без преобразований
            return cell
        elif isinstance(cell, int):  # если целое число
            return cell  # возвращаем как есть
        elif cell is None:  # если пустая ячейка
            return ""
        return str(cell).strip()  # в остальных случаях, просто преобразуем в строку

    def _create_or_replace_sheet(self, name: str):
        if name in self.wb.sheetnames:
            logger.warning(f"⚠️ Лист '{name}' уже существует. Перезапишем.")
            self.wb.remove(self.wb[name])
        else:
            logger.info(f"✅ Создан новый лист '{name}'.")
        return self.wb.create_sheet(name)

    def archive_full_report(self):
        # Проверяем существование листа Report
        if self.report_sheet not in self.wb.sheetnames:
            logger.warning(f"❌ Лист '{self.report_sheet}' не найден.")
            return

        # Получаем лист Report по имени
        source_ws = self.wb[self.report_sheet]

        # Создаем или заменяем архивный лист
        archive_ws = self._create_or_replace_sheet(self.current_month_sheet)

        # Перемещаем архивный лист сразу после "Report"
        try:
            if len(self.wb.sheetnames) > 1:
                report_index = self.wb.sheetnames.index(self.report_sheet)
                archive_index = self.wb.sheetnames.index(self.current_month_sheet)

                # Вытаскиваем объект листа
                archive_sheet = self.wb[self.current_month_sheet]

                # Удаляем из текущей позиции и вставляем после Report
                self.wb._sheets.pop(archive_index)
                self.wb._sheets.insert(report_index + 1, archive_sheet)
                logger.info(f"✅ Лист '{self.current_month_sheet}' перемещён после '{self.report_sheet}'")
        except Exception as e:
            logger.warning(f"⚠️ Не удалось переставить лист {self.current_month_sheet}: {e}")

        # Копируем данные из исходного листа
        rows_copied = 0

        # Получаем максимальный размер данных на исходном листе
        max_row = source_ws.max_row
        max_col = source_ws.max_column

        # Копируем ширину столбцов
        for col_idx in range(1, max_col + 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            if col_letter in source_ws.column_dimensions:
                archive_ws.column_dimensions[col_letter].width = source_ws.column_dimensions[col_letter].width

        # Сохраняем заголовки до удаления листа Report
        headers = []
        try:
            if max_row > 0:
                first_row = next(source_ws.iter_rows(min_row=1, max_row=1))
                headers = [cell.value for cell in first_row]
        except StopIteration:
            logger.warning("⚠️ Лист Report пуст, создаем пустой список заголовков.")

        # Копируем все значения ячеек сначала
        for row_idx in range(1, max_row + 1):
            for col_idx in range(1, max_col + 1):
                # Получаем ячейку из исходного листа
                source_cell = source_ws.cell(row=row_idx, column=col_idx)
                # Создаем ячейку в целевом листе с тем же значением
                target_cell = archive_ws.cell(row=row_idx, column=col_idx)
                # Копируем значение как есть
                target_cell.value = source_cell.value

                # Копируем формат числа
                try:
                    target_cell.number_format = source_cell.number_format

                    # Копируем стили ячейки, создавая новые объекты вместо прямого присваивания
                    if source_cell.has_style:
                        # 1. Копируем шрифт
                        if source_cell.font:
                            from openpyxl.styles import Font
                            target_cell.font = Font(
                                name=source_cell.font.name,
                                size=source_cell.font.size,
                                bold=source_cell.font.bold,
                                italic=source_cell.font.italic,
                                vertAlign=source_cell.font.vertAlign,
                                underline=source_cell.font.underline,
                                strike=source_cell.font.strike,
                                color=source_cell.font.color
                            )

                        # 2. Копируем выравнивание
                        if source_cell.alignment:
                            from openpyxl.styles import Alignment
                            target_cell.alignment = Alignment(
                                horizontal=source_cell.alignment.horizontal,
                                vertical=source_cell.alignment.vertical,
                                textRotation=source_cell.alignment.textRotation,
                                wrapText=source_cell.alignment.wrapText,
                                shrinkToFit=source_cell.alignment.shrinkToFit,
                                indent=source_cell.alignment.indent
                            )

                        # 3. Копируем границы (более сложный случай, копируем только стиль границы)
                        if source_cell.border:
                            from openpyxl.styles import Border, Side
                            sides = {}
                            for side_name in ['left', 'right', 'top', 'bottom']:
                                source_side = getattr(source_cell.border, side_name)
                                if source_side and source_side.style:
                                    sides[side_name] = Side(style=source_side.style, color=source_side.color)
                                else:
                                    sides[side_name] = Side(style=None)

                            target_cell.border = Border(**sides)

                        # 4. Копируем заливку (только если это PatternFill)
                        if source_cell.fill:
                            from openpyxl.styles import PatternFill
                            try:
                                if hasattr(source_cell.fill, 'fill_type') and source_cell.fill.fill_type:
                                    target_cell.fill = PatternFill(
                                        fill_type=source_cell.fill.fill_type,
                                        start_color=source_cell.fill.start_color,
                                        end_color=source_cell.fill.end_color
                                    )
                            except Exception:
                                # Пропускаем сложные типы заливки
                                pass

                except Exception as e:
                    logger.warning(f"⚠️ Не удалось скопировать формат ячейки {row_idx}:{col_idx}: {e}")

            rows_copied += 1

        logger.info(f"📥 Перенесено {rows_copied} строк из '{self.report_sheet}' в '{self.current_month_sheet}'.")

        # Удаляем исходный лист Report по имени
        if self.report_sheet in self.wb.sheetnames:
            self.wb.remove(self.wb[self.report_sheet])
            logger.info(f"🗑️ Лист '{self.report_sheet}' удален.")

        # Создаем новый лист Report
        new_report_ws = self.wb.create_sheet(self.report_sheet)

        # Добавляем заголовки, если они есть
        if headers:
            new_report_ws.append(headers)
            logger.info(f"📝 Заголовки добавлены в новый лист '{self.report_sheet}'.")

        # Перемещаем Report в начало книги
        try:
            # Получаем индекс нового листа Report
            new_index = self.wb.sheetnames.index(self.report_sheet)
            # Перемещаем его в начало, если он не уже там
            if new_index > 0:
                # Получаем объект листа
                sheet_to_move = self.wb[self.report_sheet]
                # Удаляем его из текущей позиции
                self.wb._sheets.pop(new_index)
                # Вставляем в начало
                self.wb._sheets.insert(0, sheet_to_move)
                logger.info(f"🔄 Лист '{self.report_sheet}' перемещён в начало.")
        except Exception as e:
            logger.warning(f"⚠️ Не удалось переместить лист '{self.report_sheet}' в начало: {e}")

        # Сохраняем изменения
        self._save()

    def _save(self):
        try:
            self.wb.save(self.filepath)
            logger.info(f"💾 Изменения сохранены в файле '{self.filepath}'.")
        except Exception as e:
            logger.error(f"❌ Ошибка при сохранении файла: {e}")